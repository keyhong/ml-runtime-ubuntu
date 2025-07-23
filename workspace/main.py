
import os
import time
from datetime import datetime
from pathlib import Path

import pandas as pd
import requests
from openpyxl.styles import Border, Font, PatternFill, Side
from tqdm import tqdm



def main():


    current_ym = datetime.now().strftime("%Y%m")  # 예: "1430"
    log_filename = f"{datetime.now().strftime('%Y-%m-%d-%a')}.xlsx"

    os.makedirs(
        Path(__file__).resolve().parents[0] / "report" / current_ym, exist_ok=True
    )

    # Specify the Excel writer and use openpyxl engine
    with pd.ExcelWriter(
        Path(__file__).resolve().parents[0]
        / "report"
        / current_ym
        / log_filename,
        engine="openpyxl",
    ) as writer:
        result_df.to_excel(writer, index=False, sheet_name="종목요약")

        # Access the workbook and the sheet
        workbook = writer.book
        sheet = workbook["종목요약"]

        # Access the header row (first row)
        header_row = sheet[1]

        # Change the header color (for example, making it light blue)
        for cell in header_row:
            cell.fill = PatternFill(
                start_color="B0E0E6", end_color="B0E0E6", fill_type="solid"
            )

        # 전체 격자에 테두리 추가
        for row in sheet.iter_rows(
            min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column
        ):
            for cell in row:
                cell.border = Border(
                    left=Side(border_style="thin", color="000000"),
                    right=Side(border_style="thin", color="000000"),
                    top=Side(border_style="thin", color="000000"),
                    bottom=Side(border_style="thin", color="000000"),
                )

        # 워크시트에서 '현재가_최고가_비율'과 '현재가_최저가_비율' 열에 대해서 텍스트 색상 적용
        for row in sheet.iter_rows(min_row=2):  # 두 번째 행부터 시작
            for cell in row:
                if (
                    isinstance(cell.value, str) and "%" in cell.value
                ):  # 값이 문자열이고 '%'이 포함된 경우
                    if cell.value.startswith("-"):
                        cell.font = Font(color="0000FF")  # '-'로 시작하면 파란색 텍스트
                    elif cell.value.startswith("+"):
                        cell.font = Font(color="FF0000")  # '+'로 시작하면 빨간색 텍스트

                if isinstance(cell.value, str) and cell.value == "신고가":
                    cell.font = Font(color="FF0000")

        # 열 너비 자동 조정
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter

            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except ValueError as _:
                    pass

            adjusted_width = max_length + 15
            sheet.column_dimensions[column].width = adjusted_width


if __name__ == "__main__":
    main()
